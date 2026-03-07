import openpyxl
class ExcelUtil():
    def __init__(self, excelPath, sheetName):
        self.data = openpyxl.load_workbook(excelPath)
        self.table = self.data[sheetName]
        # Get the total number of rows
        self.rowNum = self.table.max_row
        # Get the total number of columns
        self.colNum = self.table.max_column
        # Get the first row as the key value
        self.keys = []
        for col in range(1, self.colNum + 1):
            self.keys.append(self.table.cell(row=1, column=col).value)

    def dict_data(self):
        if self.rowNum <= 1:
            print("总行数小于1")
        else:
            r = []
            # 从第二行开始
            for row in range(2, self.rowNum + 1):
                s = {}
                for col in range(1, self.colNum + 1):
                    key = self.keys[col - 1]
                    value = self.table.cell(row=row, column=col).value
                    s[key] = value
                r.append(s)
            return r

# if __name__ == "__main__":
#     filepath = "F:\\Git\\migu-web\\data\\testdata\datas.xlsx"
#     sheetName = "a1_sta"
#     data = ExcelUtil(filepath, sheetName)
# print(data.dict_data())